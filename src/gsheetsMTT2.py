from tkinter.messagebox import *
from tkinter.filedialog import askopenfilename
from gsheets import is_iban
from decimal import Decimal, getcontext
import openpyxl


class GSheetMTT2:
    def __init__(self, _kursArg, _stdBetrag, _stdZweck):
        # diese Felder brauchen wir für den Einzug
        self.ebicsnames = ebicsnames = ["Vorname", "Name", "Kontoinhaber", "IBAN", "Preis", "Datum"]
        self.vorname = ebicsnames[0]
        self.name = ebicsnames[1]
        self.ktoinh = ebicsnames[2]
        self.iban = ebicsnames[3]
        self.betrag = ebicsnames[4]
        self.datum = ebicsnames[5]
        self.nr_einzuziehen = 0
        self.zweck = "zweck"

    @classmethod
    def getDefaults(cls):
        return "je nach Reise", "ADFC Mehrtagestouren", "je nach Reise"

    def fillEingezogen(self):
        pass

    def getStatistics(self):
        return -1, -1, -1, -1, self.nr_einzuziehen

    def getEntries(self):
        path = askopenfilename(title="Excel-Datei (edoobox) auswählen", defaultextension=".xlsx",
                               filetypes=[("XLSX", ".xlsx")])
        pathElems = path.split("_")
        entries = []
        if len(pathElems) != 4:
            showerror("", "Dateiname muß 3 Unterstriche enthalten")
            return entries
        if not pathElems[0].endswith("Export2"):
            showerror("", "Dateiname muß mit Export2 anfangen")
            return entries
        self.reisenName = pathElems[1]
        self.reisenNummer = pathElems[2]
        self.mandat = "ADFC-Mchn-MTT-" + str(self.reisenNummer)
        wb = openpyxl.load_workbook(filename=path)
        sheetnames = wb.get_sheet_names()
        if "Export" not in sheetnames:
            showerror("Ungültige Datei", "Kein Tabellenblatt 'Export' in Excel-Datei")
            return entries
        ws = wb.get_sheet_by_name("Export")
        headerMap = {}
        for (i, row) in enumerate(ws.rows):
            if i == 0:
                row0 = row
                for (j, v) in enumerate(row0):
                    headerMap[v.value] = j
                print("hdr", headerMap)
                for h in self.ebicsnames:
                    if headerMap.get(h) is None:
                        showerror("Ungültige Datei", "Tabellenblatt 'Export' hat ungültige Headerzeile, es fehlt " + h)
                        return entries
                continue
            buchung = {x: row[headerMap[x]].value for x in self.ebicsnames}
            if buchung[self.vorname] is None:
                buchung[self.vorname] = ""
            if buchung[self.name] is None:
                buchung[self.name] = ""
            if buchung[self.ktoinh] is None or buchung[self.ktoinh] == "":
                buchung[self.ktoinh] = buchung[self.vorname] + " " + buchung[self.name]
            if buchung[self.betrag] is None or \
                    buchung[self.iban] is None:
                continue
            if buchung[self.ktoinh] == "" or \
                    buchung[self.betrag] == "" or \
                    buchung[self.iban] == "":
                continue  # skip empty rows
            iban = buchung[self.iban].upper().replace(" ","");
            buchung[self.iban] = iban
            if not is_iban(iban):
                showerror("IBAN " + iban + " von " + buchung[self.ktoinh] + " ist ungültig")
                continue
            # if not is_iban(iban):
            #     showerror("Falsche IBAN", "IBAN " + iban + " von " + buchung[self.ktoinh] + " ist ungültig")
            #     continue
            betrag = buchung[self.betrag]
            if betrag >= 0:
                showerror("Betrag ungültig", "Betrag " + str(betrag) + " ist nicht negativ")
                continue
            buchung[self.betrag] = Decimal(-betrag)
            buchung[self.zweck] = "ADFC Mehrtagestour " + self.reisenName.replace("-", " ")
            buchung["mandat"] = self.mandat
            entries.append(buchung)
            self.nr_einzuziehen += 1
        return entries
